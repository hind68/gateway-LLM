from openpyxl import load_workbook

from app.ingestion.structured_routing import classify_columns


def extract_text_from_xlsx(path: str) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    parts = []
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(", ".join(cells))
    finally:
        wb.close()
    return "\n".join(parts)


def extract_xlsx_segments(path: str) -> tuple[str, str]:
    """
    Like extract_text_from_xlsx, but splits cells into (known_text,
    free_text) by column header - see structured_routing.py. Each sheet
    is classified using its own first row, since different sheets in the
    same workbook can have different columns.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    known_rows, free_rows = [], []
    try:
        for sheet in wb.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                continue
            column_is_known = classify_columns([str(h) if h is not None else "" for h in headers])

            for row in rows:
                known_cells = [
                    str(c) for i, c in enumerate(row)
                    if c is not None and i < len(column_is_known) and column_is_known[i]
                ]
                free_cells = [
                    str(c) for i, c in enumerate(row)
                    if c is not None and (i >= len(column_is_known) or not column_is_known[i])
                ]
                if known_cells:
                    known_rows.append(", ".join(known_cells))
                if free_cells:
                    free_rows.append(", ".join(free_cells))
    finally:
        wb.close()

    return "\n".join(known_rows), "\n".join(free_rows)
